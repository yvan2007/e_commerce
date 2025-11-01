from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse, HttpResponse
from django.views.generic import ListView, DetailView, CreateView, UpdateView
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.urls import reverse_lazy
from datetime import datetime, timedelta
import csv
import openpyxl
from io import BytesIO
import json
import logging

from .models import (
    Supplier, StockMovement, StockAlert, InventoryTransaction,
    InventoryReport, ProductSupplier
)
from products.models import Product

logger = logging.getLogger(__name__)


class SupplierListView(ListView):
    """Vue pour lister les fournisseurs"""
    model = Supplier
    template_name = 'inventory/supplier_list.html'
    context_object_name = 'suppliers'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Supplier.objects.all()
        
        # Filtrage par nom
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | 
                Q(contact_person__icontains=search) |
                Q(email__icontains=search)
            )
        
        return queryset.order_by('name')


class SupplierDetailView(DetailView):
    """Vue pour afficher les détails d'un fournisseur"""
    model = Supplier
    template_name = 'inventory/supplier_detail.html'
    context_object_name = 'supplier'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Produits fournis par ce fournisseur
        context['products'] = Product.objects.filter(
            productsupplier__supplier=self.object
        ).distinct()
        
        # Mouvements de stock récents
        context['recent_movements'] = StockMovement.objects.filter(
            supplier=self.object
        ).order_by('-created_at')[:10]
        
        return context


class StockMovementListView(ListView):
    """Vue pour lister les mouvements de stock"""
    model = StockMovement
    template_name = 'inventory/stock_movements.html'
    context_object_name = 'movements'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = StockMovement.objects.select_related(
            'product', 'supplier', 'user'
        ).order_by('-created_at')
        
        # Filtrage par type de mouvement
        movement_type = self.request.GET.get('type')
        if movement_type:
            queryset = queryset.filter(movement_type=movement_type)
        
        # Filtrage par produit
        product_id = self.request.GET.get('product')
        if product_id:
            queryset = queryset.filter(product_id=product_id)
        
        # Filtrage par fournisseur
        supplier_id = self.request.GET.get('supplier')
        if supplier_id:
            queryset = queryset.filter(supplier_id=supplier_id)
        
        return queryset


class StockAlertListView(ListView):
    """Vue pour lister les alertes de stock"""
    model = StockAlert
    template_name = 'inventory/stock_alerts.html'
    context_object_name = 'alerts'
    paginate_by = 20
    
    def get_queryset(self):
        return StockAlert.objects.filter(
            is_active=True
        ).select_related('product').order_by('-created_at')


class InventoryReportView(ListView):
    """Vue pour les rapports d'inventaire"""
    model = InventoryReport
    template_name = 'inventory/inventory_report.html'
    context_object_name = 'reports'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Statistiques générales
        context['total_products'] = Product.objects.count()
        context['low_stock_products'] = Product.objects.filter(
            stock_quantity__lt=10
        ).count()
        context['out_of_stock_products'] = Product.objects.filter(
            stock_quantity=0
        ).count()
        
        # Mouvements récents
        context['recent_movements'] = StockMovement.objects.order_by('-created_at')[:10]
        
        # Alertes actives
        context['active_alerts'] = StockAlert.objects.filter(
            is_active=True
        ).count()
        
        return context


class ProductStockUpdateView(UpdateView):
    """Vue pour mettre à jour le stock d'un produit"""
    model = Product
    fields = ['stock_quantity']
    template_name = 'inventory/product_stock_update.html'
    success_url = reverse_lazy('inventory:stock_movements')
    
    def form_valid(self, form):
        # Enregistrer l'ancien stock
        old_stock = self.object.stock_quantity
        
        # Mettre à jour le stock
        response = super().form_valid(form)
        
        # Créer un mouvement de stock
        new_stock = form.cleaned_data['stock_quantity']
        quantity_change = new_stock - old_stock
        
        if quantity_change != 0:
            StockMovement.objects.create(
                product=self.object,
                movement_type='adjustment',
                quantity=abs(quantity_change),
                is_incoming=quantity_change > 0,
                reason=f'Ajustement manuel: {old_stock} → {new_stock}',
                user=self.request.user
            )
        
        messages.success(
            self.request, 
            f'Stock mis à jour: {old_stock} → {new_stock}'
        )
        
        return response


@staff_member_required
def export_inventory_csv(request):
    """Exporter l'inventaire en CSV"""
    try:
        # Récupérer les paramètres
        report_type = request.GET.get('type', 'products')
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="inventory_{report_type}.csv"'
        
        writer = csv.writer(response)
        
        if report_type == 'products':
            # Export des produits
            writer.writerow(['Nom', 'SKU', 'Stock Actuel', 'Prix', 'Catégorie', 'Statut'])
            
            products = Product.objects.select_related('category')
            for product in products:
                writer.writerow([
                    product.name,
                    product.sku,
                    product.stock_quantity,
                    product.price,
                    product.category.name if product.category else '',
                    'En Stock' if product.stock_quantity > 0 else 'Rupture'
                ])
        
        elif report_type == 'movements':
            # Export des mouvements
            writer.writerow(['Date', 'Produit', 'Type', 'Quantité', 'Fournisseur', 'Utilisateur'])
            
            movements = StockMovement.objects.select_related(
                'product', 'supplier', 'user'
            ).order_by('-created_at')
            
            for movement in movements:
                writer.writerow([
                    movement.created_at.strftime('%Y-%m-%d %H:%M'),
                    movement.product.name,
                    movement.get_movement_type_display(),
                    movement.quantity,
                    movement.supplier.name if movement.supplier else '',
                    movement.user.email if movement.user else ''
                ])
        
        elif report_type == 'suppliers':
            # Export des fournisseurs
            writer.writerow(['Nom', 'Contact', 'Email', 'Téléphone', 'Adresse', 'Produits'])
            
            suppliers = Supplier.objects.all()
            for supplier in suppliers:
                product_count = Product.objects.filter(
                    productsupplier__supplier=supplier
                ).count()
                
                writer.writerow([
                    supplier.name,
                    supplier.contact_person,
                    supplier.email,
                    supplier.phone,
                    supplier.address,
                    product_count
                ])
        
        return response
        
    except Exception as e:
        logger.error(f"Erreur export CSV: {e}")
        messages.error(request, "Erreur lors de l'export CSV.")
        return redirect('inventory:inventory_report')


@staff_member_required
def export_inventory_excel(request):
    """Exporter l'inventaire en Excel"""
    try:
        # Récupérer les paramètres
        report_type = request.GET.get('type', 'products')
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="inventory_{report_type}.xlsx"'
        
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Inventaire {report_type.title()}"
        
        if report_type == 'products':
            # Headers
            headers = ['Nom', 'SKU', 'Stock Actuel', 'Prix', 'Catégorie', 'Statut']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Données
            products = Product.objects.select_related('category')
            for row, product in enumerate(products, 2):
                ws.cell(row=row, column=1, value=product.name)
                ws.cell(row=row, column=2, value=product.sku)
                ws.cell(row=row, column=3, value=product.stock_quantity)
                ws.cell(row=row, column=4, value=product.price)
                ws.cell(row=row, column=5, value=product.category.name if product.category else '')
                ws.cell(row=row, column=6, value='En Stock' if product.stock_quantity > 0 else 'Rupture')
        
        elif report_type == 'movements':
            # Headers
            headers = ['Date', 'Produit', 'Type', 'Quantité', 'Fournisseur', 'Utilisateur']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Données
            movements = StockMovement.objects.select_related(
                'product', 'supplier', 'user'
            ).order_by('-created_at')
            
            for row, movement in enumerate(movements, 2):
                ws.cell(row=row, column=1, value=movement.created_at.strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row, column=2, value=movement.product.name)
                ws.cell(row=row, column=3, value=movement.get_movement_type_display())
                ws.cell(row=row, column=4, value=movement.quantity)
                ws.cell(row=row, column=5, value=movement.supplier.name if movement.supplier else '')
                ws.cell(row=row, column=6, value=movement.user.email if movement.user else '')
        
        # Sauvegarder dans la réponse
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Erreur export Excel: {e}")
        messages.error(request, "Erreur lors de l'export Excel.")
        return redirect('inventory:inventory_report')


@login_required
def low_stock_alert(request):
    """API pour vérifier les alertes de stock bas"""
    try:
        # Produits avec stock bas
        low_stock_products = Product.objects.filter(
            stock_quantity__lt=10,
            stock_quantity__gt=0
        ).values('id', 'name', 'sku', 'stock_quantity')
        
        # Produits en rupture
        out_of_stock_products = Product.objects.filter(
            stock_quantity=0
        ).values('id', 'name', 'sku', 'stock_quantity')
        
        return JsonResponse({
            'low_stock': list(low_stock_products),
            'out_of_stock': list(out_of_stock_products),
            'total_low_stock': low_stock_products.count(),
            'total_out_of_stock': out_of_stock_products.count()
        })
        
    except Exception as e:
        logger.error(f"Erreur low stock alert: {e}")
        return JsonResponse({'error': 'Erreur lors de la vérification'}, status=500)


@staff_member_required
def bulk_stock_update(request):
    """Mise à jour en masse du stock"""
    if request.method == 'POST':
        try:
            # Récupérer les données JSON
            data = json.loads(request.body)
            updates = data.get('updates', [])
            
            updated_count = 0
            
            for update in updates:
                product_id = update.get('product_id')
                new_stock = update.get('stock_quantity')
                reason = update.get('reason', 'Mise à jour en masse')
                
                if product_id and new_stock is not None:
                    product = Product.objects.get(id=product_id)
                    old_stock = product.stock_quantity
                    
                    # Mettre à jour le stock
                    product.stock_quantity = new_stock
                    product.save()
                    
                    # Créer un mouvement
                    quantity_change = new_stock - old_stock
                    if quantity_change != 0:
                        StockMovement.objects.create(
                            product=product,
                            movement_type='adjustment',
                            quantity=abs(quantity_change),
                            is_incoming=quantity_change > 0,
                            reason=f'{reason}: {old_stock} → {new_stock}',
                            user=request.user
                        )
                    
                    updated_count += 1
            
            return JsonResponse({
                'success': True,
                'message': f'{updated_count} produits mis à jour avec succès'
            })
            
        except Exception as e:
            logger.error(f"Erreur bulk stock update: {e}")
            return JsonResponse({
                'success': False,
                'message': 'Erreur lors de la mise à jour'
            }, status=500)
    
    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)